import streamlit as st
import pandas as pd
import sqlite3
import plotly.express as px
from io import BytesIO
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference

st.set_page_config(page_title="Budget & Expense Analyzer", layout="wide")

st.title("Budget & Expense Analyzer with SQL")

DB_NAME = "budget.db"


# -----------------------------
# Database Functions
# -----------------------------
def create_database():
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS transactions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            transaction_date TEXT,
            description TEXT,
            amount REAL,
            transaction_type TEXT,
            category TEXT,
            month TEXT
        )
    """)

    conn.commit()
    conn.close()


def replace_transactions(df):
    conn = sqlite3.connect(DB_NAME)

    conn.execute("DELETE FROM transactions")

    df.to_sql(
        "transactions",
        conn,
        if_exists="append",
        index=False
    )

    conn.commit()
    conn.close()


def run_query(query):
    conn = sqlite3.connect(DB_NAME)
    result = pd.read_sql_query(query, conn)
    conn.close()
    return result


# -----------------------------
# Data Cleaning Function
# -----------------------------
def clean_transactions(df):
    df = df.copy()
    df.columns = df.columns.str.strip()

    required_columns = ["Date", "Description", "Amount", "Category"]

    for column in required_columns:
        if column not in df.columns:
            raise ValueError(f"Missing required column: {column}")

    df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
    df["Amount"] = pd.to_numeric(df["Amount"], errors="coerce")

    df = df.dropna(subset=["Date", "Amount"])

    df["transaction_type"] = df["Amount"].apply(
        lambda amount: "Income" if amount >= 0 else "Expense"
    )

    df["amount"] = df["Amount"].abs()
    df["transaction_date"] = df["Date"].dt.strftime("%Y-%m-%d")
    df["description"] = df["Description"]
    df["category"] = df["Category"]
    df["month"] = df["Date"].dt.strftime("%Y-%m")

    cleaned_df = df[
        [
            "transaction_date",
            "description",
            "amount",
            "transaction_type",
            "category",
            "month",
        ]
    ]

    return cleaned_df


# -----------------------------
# Create Database
# -----------------------------
create_database()


# -----------------------------
# Tabs
# -----------------------------
upload_tab, dashboard_tab, sql_tab, categories_tab, trends_tab, export_tab = st.tabs([
    "Upload",
    "Dashboard",
    "SQL Analysis",
    "Categories",
    "Trends",
    "Export"
])


# -----------------------------
# Upload Tab
# -----------------------------
with upload_tab:
    st.subheader("Upload Transactions CSV")

    uploaded_file = st.file_uploader("Upload a CSV file", type=["csv"])

    if uploaded_file is not None:
        original_df = pd.read_csv(uploaded_file)

        st.write("Original Uploaded Data")
        st.dataframe(original_df, use_container_width=True)

        try:
            cleaned_df = clean_transactions(original_df)

            st.write("Cleaned Data")
            st.dataframe(cleaned_df, use_container_width=True)

            replace_transactions(cleaned_df)

            st.success("Transactions automatically saved to SQLite database.")

        except Exception as error:
            st.error(f"Error cleaning file: {error}")


# -----------------------------
# Dashboard Tab
# -----------------------------
with dashboard_tab:
    st.subheader("Financial Summary")

    total_income_df = run_query("""
        SELECT COALESCE(SUM(amount), 0) AS total_income
        FROM transactions
        WHERE transaction_type = 'Income'
    """)

    total_expenses_df = run_query("""
        SELECT COALESCE(SUM(amount), 0) AS total_expenses
        FROM transactions
        WHERE transaction_type = 'Expense'
    """)

    total_income = float(total_income_df["total_income"].iloc[0])
    total_expenses = float(total_expenses_df["total_expenses"].iloc[0])
    net_savings = total_income - total_expenses
    savings_rate = net_savings / total_income if total_income > 0 else 0

    col1, col2, col3, col4 = st.columns(4)

    col1.metric("Total Income", f"${total_income:,.2f}")
    col2.metric("Total Expenses", f"${total_expenses:,.2f}")
    col3.metric("Net Savings", f"${net_savings:,.2f}")
    col4.metric("Savings Rate", f"{savings_rate:.2%}")

    st.subheader("All Transactions from SQL")

    all_transactions_df = run_query("""
        SELECT *
        FROM transactions
        ORDER BY transaction_date
    """)

    st.dataframe(all_transactions_df, use_container_width=True)


# -----------------------------
# SQL Analysis Tab
# -----------------------------
with sql_tab:
    st.subheader("SQL Analysis")

    st.write("Total Income Query")
    st.code("""
SELECT SUM(amount) AS total_income
FROM transactions
WHERE transaction_type = 'Income';
""", language="sql")
    st.dataframe(total_income_df, use_container_width=True)

    st.write("Total Expenses Query")
    st.code("""
SELECT SUM(amount) AS total_expenses
FROM transactions
WHERE transaction_type = 'Expense';
""", language="sql")
    st.dataframe(total_expenses_df, use_container_width=True)

    st.write("Top 10 Largest Expenses")

    top_expenses_query = """
        SELECT transaction_date, description, category, amount
        FROM transactions
        WHERE transaction_type = 'Expense'
        ORDER BY amount DESC
        LIMIT 10
    """

    st.code(top_expenses_query, language="sql")
    top_expenses_df = run_query(top_expenses_query)
    st.dataframe(top_expenses_df, use_container_width=True)


# -----------------------------
# Categories Tab
# -----------------------------
with categories_tab:
    st.subheader("Spending by Category")

    category_query = """
        SELECT category, SUM(amount) AS total_spent
        FROM transactions
        WHERE transaction_type = 'Expense'
        GROUP BY category
        ORDER BY total_spent DESC
    """

    category_df = run_query(category_query)

    if not category_df.empty:
        st.dataframe(category_df, use_container_width=True)

        fig = px.bar(
            category_df,
            x="category",
            y="total_spent",
            title="Expenses by Category",
            labels={
                "category": "Category",
                "total_spent": "Total Spent"
            }
        )

        st.plotly_chart(fig, use_container_width=True)
    else:
        st.info("No expense data found yet. Upload and save transactions first.")


# -----------------------------
# Trends Tab
# -----------------------------
with trends_tab:
    st.subheader("Monthly Income vs Expenses")

    monthly_query = """
        SELECT 
            month,
            SUM(CASE WHEN transaction_type = 'Income' THEN amount ELSE 0 END) AS income,
            SUM(CASE WHEN transaction_type = 'Expense' THEN amount ELSE 0 END) AS expenses
        FROM transactions
        GROUP BY month
        ORDER BY month
    """

    monthly_df = run_query(monthly_query)

    if not monthly_df.empty:
        monthly_df["net_savings"] = monthly_df["income"] - \
            monthly_df["expenses"]

        st.dataframe(monthly_df, use_container_width=True)

        monthly_long_df = monthly_df.melt(
            id_vars="month",
            value_vars=["income", "expenses"],
            var_name="Type",
            value_name="Amount"
        )

        monthly_bar = px.bar(
            monthly_long_df,
            x="month",
            y="Amount",
            color="Type",
            barmode="group",
            title="Monthly Income vs Expenses",
            labels={
                "month": "Month",
                "Amount": "Amount"
            }
        )

        st.plotly_chart(monthly_bar, use_container_width=True)

        savings_line = px.line(
            monthly_df,
            x="month",
            y="net_savings",
            markers=True,
            title="Monthly Net Savings",
            labels={
                "month": "Month",
                "net_savings": "Net Savings"
            }
        )

        st.plotly_chart(savings_line, use_container_width=True)

    else:
        st.info("No monthly data found yet. Upload and save transactions first.")

    st.subheader("Daily Expense Trend")

    daily_expense_query = """
        SELECT 
            transaction_date,
            SUM(amount) AS daily_expenses
        FROM transactions
        WHERE transaction_type = 'Expense'
        GROUP BY transaction_date
        ORDER BY transaction_date
    """

    daily_expense_df = run_query(daily_expense_query)

    if not daily_expense_df.empty:
        st.dataframe(daily_expense_df, use_container_width=True)

        daily_expense_chart = px.line(
            daily_expense_df,
            x="transaction_date",
            y="daily_expenses",
            markers=True,
            title="Daily Expense Trend",
            labels={
                "transaction_date": "Date",
                "daily_expenses": "Daily Expenses"
            }
        )

        st.plotly_chart(daily_expense_chart, use_container_width=True)
    else:
        st.info("No daily expense data found yet.")
# -----------------------------
# Export Tab
# -----------------------------
with export_tab:
    st.subheader("Export Professional Budget Report")

    all_transactions_export = run_query("""
        SELECT *
        FROM transactions
        ORDER BY transaction_date
    """)

    monthly_summary_export = run_query("""
        SELECT 
            month,
            SUM(CASE WHEN transaction_type = 'Income' THEN amount ELSE 0 END) AS income,
            SUM(CASE WHEN transaction_type = 'Expense' THEN amount ELSE 0 END) AS expenses,
            SUM(CASE WHEN transaction_type = 'Income' THEN amount ELSE 0 END) -
            SUM(CASE WHEN transaction_type = 'Expense' THEN amount ELSE 0 END) AS net_savings
        FROM transactions
        GROUP BY month
        ORDER BY month
    """)

    category_summary_export = run_query("""
        SELECT category, SUM(amount) AS total_spent
        FROM transactions
        WHERE transaction_type = 'Expense'
        GROUP BY category
        ORDER BY total_spent DESC
    """)

    top_expenses_export = run_query("""
        SELECT transaction_date, description, category, amount
        FROM transactions
        WHERE transaction_type = 'Expense'
        ORDER BY amount DESC
        LIMIT 10
    """)

    daily_expenses_export = run_query("""
        SELECT 
            transaction_date,
            SUM(amount) AS daily_expenses
        FROM transactions
        WHERE transaction_type = 'Expense'
        GROUP BY transaction_date
        ORDER BY transaction_date
    """)

    if all_transactions_export.empty:
        st.info("No transactions found yet. Upload and save transactions first.")

    else:
        buffer = BytesIO()

        total_income = monthly_summary_export["income"].sum()
        total_expenses = monthly_summary_export["expenses"].sum()
        net_savings = total_income - total_expenses
        savings_rate = net_savings / total_income if total_income > 0 else 0

        largest_category = "N/A"
        largest_category_amount = 0

        if not category_summary_export.empty:
            largest_category = category_summary_export.iloc[0]["category"]
            largest_category_amount = category_summary_export.iloc[0]["total_spent"]

        largest_expense_description = "N/A"
        largest_expense_amount = 0

        if not top_expenses_export.empty:
            largest_expense_description = top_expenses_export.iloc[0]["description"]
            largest_expense_amount = top_expenses_export.iloc[0]["amount"]

        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            all_transactions_export.to_excel(
                writer, sheet_name="All Transactions", index=False)
            monthly_summary_export.to_excel(
                writer, sheet_name="Monthly Summary", index=False)
            category_summary_export.to_excel(
                writer, sheet_name="Category Summary", index=False)
            top_expenses_export.to_excel(
                writer, sheet_name="Top Expenses", index=False)
            daily_expenses_export.to_excel(
                writer, sheet_name="Daily Expense Trend", index=False)

            workbook = writer.book

            # -----------------------------
            # Professional Styling Settings
            # -----------------------------
            dark_blue_fill = PatternFill(
                start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            medium_blue_fill = PatternFill(
                start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
            green_fill = PatternFill(
                start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
            red_fill = PatternFill(start_color="FCE4D6",
                                   end_color="FCE4D6", fill_type="solid")

            white_font = Font(color="FFFFFF", bold=True)
            title_font = Font(size=16, bold=True, color="1F4E78")
            bold_font = Font(bold=True)

            thin_border = Border(
                left=Side(style="thin", color="D9E2F3"),
                right=Side(style="thin", color="D9E2F3"),
                top=Side(style="thin", color="D9E2F3"),
                bottom=Side(style="thin", color="D9E2F3")
            )

            currency_format = '$#,##0.00'
            percent_format = '0.00%'

            # -----------------------------
            # Create Executive Summary Sheet
            # -----------------------------
            summary_ws = workbook.create_sheet("Executive Summary", 0)

            summary_ws["A1"] = "Budget & Expense Analyzer Report"
            summary_ws["A1"].font = title_font
            summary_ws.merge_cells("A1:D1")

            summary_ws["A2"] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %I:%M %p')}"
            summary_ws["A2"].font = Font(italic=True, color="666666")

            summary_data = [
                ["Metric", "Value"],
                ["Total Income", total_income],
                ["Total Expenses", total_expenses],
                ["Net Savings", net_savings],
                ["Savings Rate", savings_rate],
                ["Largest Spending Category", largest_category],
                ["Largest Category Amount", largest_category_amount],
                ["Largest Expense", largest_expense_description],
                ["Largest Expense Amount", largest_expense_amount],
            ]

            start_row = 4

            for row_index, row_data in enumerate(summary_data, start=start_row):
                for col_index, value in enumerate(row_data, start=1):
                    cell = summary_ws.cell(
                        row=row_index, column=col_index, value=value)
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center")

                    if row_index == start_row:
                        cell.fill = dark_blue_fill
                        cell.font = white_font
                    elif col_index == 1:
                        cell.fill = medium_blue_fill
                        cell.font = bold_font

            for row in range(start_row + 1, start_row + len(summary_data)):
                metric_name = summary_ws.cell(row=row, column=1).value
                value_cell = summary_ws.cell(row=row, column=2)

                if metric_name in [
                    "Total Income",
                    "Total Expenses",
                    "Net Savings",
                    "Largest Category Amount",
                    "Largest Expense Amount"
                ]:
                    value_cell.number_format = currency_format

                if metric_name == "Savings Rate":
                    value_cell.number_format = percent_format

            summary_ws.column_dimensions["A"].width = 28
            summary_ws.column_dimensions["B"].width = 28
            summary_ws.column_dimensions["C"].width = 18
            summary_ws.column_dimensions["D"].width = 18

            summary_ws["A15"] = "Report Sheets Included"
            summary_ws["A15"].font = bold_font
            summary_ws["A15"].fill = dark_blue_fill
            summary_ws["A15"].font = white_font

            included_sheets = [
                "All Transactions",
                "Monthly Summary",
                "Category Summary",
                "Top Expenses",
                "Daily Expense Trend",
                "Charts"
            ]

            for index, sheet_name in enumerate(included_sheets, start=16):
                summary_ws.cell(row=index, column=1, value=sheet_name)
                summary_ws.cell(row=index, column=1).border = thin_border

            # -----------------------------
            # Format Data Sheets
            # -----------------------------
            for sheet_name in [
                "All Transactions",
                "Monthly Summary",
                "Category Summary",
                "Top Expenses",
                "Daily Expense Trend"
            ]:
                worksheet = workbook[sheet_name]

                worksheet.freeze_panes = "A2"
                worksheet.auto_filter.ref = worksheet.dimensions

                for cell in worksheet[1]:
                    cell.fill = dark_blue_fill
                    cell.font = white_font
                    cell.alignment = Alignment(
                        horizontal="center", vertical="center")
                    cell.border = thin_border

                for row in worksheet.iter_rows(min_row=2):
                    for cell in row:
                        cell.border = thin_border
                        cell.alignment = Alignment(vertical="center")

                        if cell.column_letter in ["D", "B", "C"] and sheet_name in [
                            "Monthly Summary",
                            "Category Summary",
                            "Top Expenses",
                            "Daily Expense Trend"
                        ]:
                            if isinstance(cell.value, (int, float)):
                                cell.number_format = currency_format

                        if sheet_name == "All Transactions" and cell.column_letter == "D":
                            if isinstance(cell.value, (int, float)):
                                cell.number_format = currency_format

                for column_cells in worksheet.columns:
                    max_length = 0
                    column_letter = column_cells[0].column_letter

                    for cell in column_cells:
                        try:
                            max_length = max(max_length, len(str(cell.value)))
                        except Exception:
                            pass

                    worksheet.column_dimensions[column_letter].width = min(
                        max_length + 3, 35)

            # -----------------------------
            # Create Charts Sheet
            # -----------------------------
            charts_ws = workbook.create_sheet("Charts")

            charts_ws["A1"] = "Budget & Expense Charts"
            charts_ws["A1"].font = title_font
            charts_ws.merge_cells("A1:H1")

            charts_ws["A2"] = "These charts are generated directly inside Excel from SQL-based summary tables."
            charts_ws["A2"].font = Font(italic=True, color="666666")
            charts_ws.merge_cells("A2:H2")

            # Chart 1: Spending by Category
            category_ws = workbook["Category Summary"]

            if category_ws.max_row > 1:
                category_chart = BarChart()
                category_chart.type = "bar"
                category_chart.style = 10
                category_chart.title = "Spending by Category"
                category_chart.y_axis.title = "Category"
                category_chart.x_axis.title = "Total Spent"

                category_data = Reference(
                    category_ws,
                    min_col=2,
                    min_row=1,
                    max_row=category_ws.max_row
                )

                category_labels = Reference(
                    category_ws,
                    min_col=1,
                    min_row=2,
                    max_row=category_ws.max_row
                )

                category_chart.add_data(category_data, titles_from_data=True)
                category_chart.set_categories(category_labels)
                category_chart.height = 9
                category_chart.width = 18

                charts_ws.add_chart(category_chart, "A4")

            # Chart 2: Monthly Income vs Expenses
            monthly_ws = workbook["Monthly Summary"]

            if monthly_ws.max_row > 1:
                monthly_chart = BarChart()
                monthly_chart.type = "col"
                monthly_chart.style = 10
                monthly_chart.title = "Monthly Income vs Expenses"
                monthly_chart.y_axis.title = "Amount"
                monthly_chart.x_axis.title = "Month"

                monthly_data = Reference(
                    monthly_ws,
                    min_col=2,
                    max_col=3,
                    min_row=1,
                    max_row=monthly_ws.max_row
                )

                monthly_labels = Reference(
                    monthly_ws,
                    min_col=1,
                    min_row=2,
                    max_row=monthly_ws.max_row
                )

                monthly_chart.add_data(monthly_data, titles_from_data=True)
                monthly_chart.set_categories(monthly_labels)
                monthly_chart.height = 9
                monthly_chart.width = 18

                charts_ws.add_chart(monthly_chart, "A22")

            # Chart 3: Monthly Net Savings
            if monthly_ws.max_row > 1:
                savings_chart = LineChart()
                savings_chart.style = 13
                savings_chart.title = "Monthly Net Savings"
                savings_chart.y_axis.title = "Net Savings"
                savings_chart.x_axis.title = "Month"

                savings_data = Reference(
                    monthly_ws,
                    min_col=4,
                    min_row=1,
                    max_row=monthly_ws.max_row
                )

                savings_labels = Reference(
                    monthly_ws,
                    min_col=1,
                    min_row=2,
                    max_row=monthly_ws.max_row
                )

                savings_chart.add_data(savings_data, titles_from_data=True)
                savings_chart.set_categories(savings_labels)
                savings_chart.height = 9
                savings_chart.width = 18

                charts_ws.add_chart(savings_chart, "A40")

            # Chart 4: Daily Expense Trend
            daily_ws = workbook["Daily Expense Trend"]

            if daily_ws.max_row > 1:
                daily_chart = LineChart()
                daily_chart.style = 13
                daily_chart.title = "Daily Expense Trend"
                daily_chart.y_axis.title = "Daily Expenses"
                daily_chart.x_axis.title = "Date"

                daily_data = Reference(
                    daily_ws,
                    min_col=2,
                    min_row=1,
                    max_row=daily_ws.max_row
                )

                daily_labels = Reference(
                    daily_ws,
                    min_col=1,
                    min_row=2,
                    max_row=daily_ws.max_row
                )

                daily_chart.add_data(daily_data, titles_from_data=True)
                daily_chart.set_categories(daily_labels)
                daily_chart.height = 9
                daily_chart.width = 18

                charts_ws.add_chart(daily_chart, "A58")

            charts_ws.column_dimensions["A"].width = 18
            charts_ws.column_dimensions["B"].width = 18
            charts_ws.column_dimensions["C"].width = 18
            charts_ws.column_dimensions["D"].width = 18
            charts_ws.column_dimensions["E"].width = 18
            charts_ws.column_dimensions["F"].width = 18
            charts_ws.column_dimensions["G"].width = 18
            charts_ws.column_dimensions["H"].width = 18

        st.download_button(
            label="📊 Download Professional Budget Excel Report",
            data=buffer.getvalue(),
            file_name="professional_budget_expense_report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
